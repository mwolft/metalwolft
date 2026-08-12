import copy
import shutil
import sys
import unittest
import uuid
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
SRC_DIR = ROOT_DIR / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))


from api.aeat_sales_ledger_service import (  # noqa: E402
    AEAT_HEADER_ROW_1,
    AEAT_HEADER_ROW_2,
    AEAT_SALES_LEDGER_SHEET_NAME,
    DATE_FORMAT,
    MONEY_FORMAT,
    AeatSalesLedgerValidationError,
    AeatSalesLedgerWriteError,
    export_aeat_sales_ledger,
    generate_aeat_sales_ledger_workbook,
)
from api.invoice_snapshot_integrity import calculate_invoice_snapshot_hash  # noqa: E402


TEST_TMP_ROOT = ROOT_DIR / "tmp-aeat-sales-ledger-tests"


@dataclass
class InvoiceDouble:
    id: int
    invoice_number: str
    invoice_snapshot: dict | None
    invoice_snapshot_hash: str | None
    issued_at: datetime | None = datetime(2026, 7, 18, 10, 0, 0)
    invoice_type: str | None = "ordinary"
    original_invoice_id: int | None = None
    rectification_aeat_type: str | None = None
    rectification_aeat_classified_at: datetime | None = None
    rectification_aeat_classified_by: str | None = None
    original_invoice: object | None = None

    @property
    def order(self):
        raise AssertionError("AEAT ledger must not read live order data")

    @property
    def user(self):
        raise AssertionError("AEAT ledger must not read live user data")

    @property
    def checkout_session(self):
        raise AssertionError("AEAT ledger must not read live checkout data")


@dataclass
class AccountingEntryDouble:
    id: int
    invoice_date: date
    invoice_number: str
    invoice: InvoiceDouble
    entry_type: str = "sale"
    currency: str = "EUR"
    customer_name: str = "Projection Customer"
    customer_tax_id: str | None = "PROJECTION-TAX"
    taxable_base: Decimal = Decimal("100.00")
    vat_amount: Decimal = Decimal("21.00")
    total_amount: Decimal = Decimal("121.00")
    payment_provider: str | None = "stripe"
    order_id: int | None = 123
    status: str = "pending"

    @property
    def order(self):
        raise AssertionError("AEAT ledger must not read live order relationship")

    @property
    def user(self):
        raise AssertionError("AEAT ledger must not read live user relationship")

    @property
    def checkout_session(self):
        raise AssertionError("AEAT ledger must not read live checkout relationship")


@contextmanager
def temp_export_dir():
    TEST_TMP_ROOT.mkdir(exist_ok=True)
    case_dir = TEST_TMP_ROOT / f"case-{uuid.uuid4().hex}"
    case_dir.mkdir()
    try:
        yield case_dir
    finally:
        shutil.rmtree(case_dir, ignore_errors=True)


def tearDownModule():
    shutil.rmtree(TEST_TMP_ROOT, ignore_errors=True)


def snapshot(**overrides):
    data = {
        "schema_version": 1,
        "metadata": {"generator": "invoice_snapshot_builder_v1"},
        "issuer": {
            "legal_name": "MetalWolft",
            "tax_id": "B00000000",
            "country_code": "ES",
        },
        "customer": {
            "legal_name": "Cliente Fiscal",
            "tax_id": "00000000T",
            "country_code": "ES",
        },
        "operation": {
            "invoice_type": "ordinary",
            "issue_date": "2026-07-18",
            "operation_date": "2026-07-17",
            "currency": "EUR",
            "order_id": 42,
        },
        "lines": [
            {
                "line_number": 1,
                "description": "Reja a medida",
                "tax_rate": "21.00",
                "tax_base": "100.00",
                "tax_amount": "21.00",
                "line_total": "121.00",
            }
        ],
        "totals": {
            "tax_base": "100.00",
            "tax_amount": "21.00",
            "total_amount": "121.00",
        },
        "payment": {"provider": "stripe"},
        "references": {"order_id": 42},
    }
    data.update(overrides)
    return data


def invoice(
    *,
    invoice_id=1,
    invoice_number="F2026000001",
    invoice_snapshot=None,
    stored_hash=None,
    invoice_type="ordinary",
    original_invoice_id=None,
    rectification_aeat_type=None,
    rectification_aeat_classified_at=None,
    rectification_aeat_classified_by=None,
    original_invoice=None,
):
    fiscal_snapshot = snapshot() if invoice_snapshot is None else invoice_snapshot
    return InvoiceDouble(
        id=invoice_id,
        invoice_number=invoice_number,
        invoice_snapshot=fiscal_snapshot,
        invoice_snapshot_hash=stored_hash or (
            calculate_invoice_snapshot_hash(fiscal_snapshot)
            if isinstance(fiscal_snapshot, dict)
            else None
        ),
        invoice_type=invoice_type,
        original_invoice_id=original_invoice_id,
        rectification_aeat_type=rectification_aeat_type,
        rectification_aeat_classified_at=rectification_aeat_classified_at,
        rectification_aeat_classified_by=rectification_aeat_classified_by,
        original_invoice=original_invoice,
    )


def entry(**overrides):
    data = {
        "id": 1,
        "invoice_date": date(2026, 7, 18),
        "invoice_number": "F2026000001",
        "invoice": invoice(),
    }
    data.update(overrides)
    return AccountingEntryDouble(**data)


def corrective_snapshot(*, aeat_type="R1", original_invoice_id=1, original_invoice_number="F2026000001", **overrides):
    original = snapshot()
    data = {
        **copy.deepcopy(original),
        "schema_version": 3,
        "metadata": {"generator": "invoice_snapshot_builder_v3"},
        "operation": {
            **copy.deepcopy(original["operation"]),
            "invoice_type": "corrective",
            "issue_date": "2026-07-19",
            "rectification": {
                "rectification_type": "differences",
                "rectification_scope": "total",
                "rectification_reason": "invoice_error",
                "rectification_reason_text": "Factura emitida por error",
                "aeat_type": aeat_type,
                "original_invoice_id": original_invoice_id,
                "original_invoice_number": original_invoice_number,
                "original_invoice_issued_at": "2026-07-18T10:00:00",
                "affected_line_numbers": [1],
            },
        },
        "lines": [{
            **copy.deepcopy(original["lines"][0]),
            "tax_base": "-100.00",
            "tax_amount": "-21.00",
            "line_total": "-121.00",
        }],
        "totals": {
            "tax_base": "-100.00",
            "tax_amount": "-21.00",
            "total_amount": "-121.00",
        },
    }
    data.update(overrides)
    return data


def corrective_entry(*, aeat_type="R1", snapshot_overrides=None, **overrides):
    original = invoice(invoice_id=41, invoice_number="F2026000001")
    fiscal_snapshot = corrective_snapshot(
        original_invoice_id=original.id,
        original_invoice_number=original.invoice_number,
        aeat_type=aeat_type,
        **(snapshot_overrides or {}),
    )
    corrective = invoice(
        invoice_id=42,
        invoice_number="R2026000001",
        invoice_snapshot=fiscal_snapshot,
        invoice_type="corrective",
        original_invoice_id=41,
        rectification_aeat_type=aeat_type,
        original_invoice=original,
    )
    data = {
        "id": 2,
        "invoice_date": date(2026, 7, 19),
        "invoice_number": "R2026000001",
        "invoice": corrective,
        "taxable_base": Decimal("-100.00"),
        "vat_amount": Decimal("-21.00"),
        "total_amount": Decimal("-121.00"),
    }
    data.update(overrides)
    return AccountingEntryDouble(**data)


def legacy_corrective_entry(*, aeat_type=None, classified_at=None, classified_by=None, **overrides):
    accounting_entry = corrective_entry(aeat_type=aeat_type, **overrides)
    rectification = accounting_entry.invoice.invoice_snapshot["operation"]["rectification"]
    rectification.pop("aeat_type")
    accounting_entry.invoice.invoice_snapshot_hash = calculate_invoice_snapshot_hash(
        accounting_entry.invoice.invoice_snapshot
    )
    accounting_entry.invoice.rectification_aeat_classified_at = classified_at
    accounting_entry.invoice.rectification_aeat_classified_by = classified_by
    return accounting_entry


def export_and_open(entries, path):
    result = export_aeat_sales_ledger(entries, output_path=path)
    workbook = load_workbook(path, data_only=True)
    return result, workbook


def normalized_header(header):
    return [value or None for value in header]


class AeatSalesLedgerServiceTest(unittest.TestCase):
    def test_valid_export_creates_sheet_headers_and_single_row(self):
        with temp_export_dir() as tmpdir:
            result, workbook = export_and_open([entry()], tmpdir / "aeat.xlsx")
            sheet = workbook[AEAT_SALES_LEDGER_SHEET_NAME]

            self.assertEqual(result.filename, "aeat.xlsx")
            self.assertEqual(result.row_count, 1)
            self.assertGreater(result.file_size, 0)
            self.assertEqual(sheet.max_column, 36)
            self.assertEqual([cell.value for cell in sheet[1]], normalized_header(AEAT_HEADER_ROW_1))
            self.assertEqual([cell.value for cell in sheet[2]], normalized_header(AEAT_HEADER_ROW_2))

    def test_valid_export_accepts_v2_snapshot(self):
        fiscal_snapshot = snapshot(schema_version=2, metadata={"generator": "invoice_snapshot_builder_v2"})

        with temp_export_dir() as tmpdir:
            result, workbook = export_and_open([entry(invoice=invoice(invoice_snapshot=fiscal_snapshot))], tmpdir / "aeat.xlsx")

        self.assertEqual(result.row_count, 1)
        self.assertEqual(workbook[AEAT_SALES_LEDGER_SHEET_NAME]["L3"].value, "F2026000001")

    def test_classified_total_rectificatives_export_as_r1_or_r4(self):
        for aeat_type in ("R1", "R4"):
            with self.subTest(aeat_type=aeat_type), temp_export_dir() as tmpdir:
                accounting_entry = corrective_entry(aeat_type=aeat_type)
                before_snapshot = copy.deepcopy(accounting_entry.invoice.invoice_snapshot)
                result, workbook = export_and_open([accounting_entry], tmpdir / "aeat.xlsx")
                sheet = workbook[AEAT_SALES_LEDGER_SHEET_NAME]

                self.assertEqual(result.row_count, 1)
                self.assertEqual(sheet["F3"].value, aeat_type)
                self.assertEqual(sheet["L3"].value, "R2026000001")
                self.assertEqual(sheet["I3"].value.date(), date(2026, 7, 19))
                self.assertEqual(sheet["AJ3"].value, "F2026000001")
                self.assertEqual(sheet["H3"].value, Decimal("-100.00"))
                self.assertEqual(sheet["V3"].value, Decimal("-100.00"))
                self.assertEqual(sheet["X3"].value, Decimal("-21.00"))
                self.assertEqual(sheet["U3"].value, Decimal("-121.00"))
                self.assertEqual(accounting_entry.invoice.invoice_snapshot, before_snapshot)
                self.assertEqual(accounting_entry.status, "pending")

    def test_legacy_total_rectificatives_require_audited_manual_r1_or_r4(self):
        for aeat_type in ("R1", "R4"):
            with self.subTest(aeat_type=aeat_type), temp_export_dir() as tmpdir:
                accounting_entry = legacy_corrective_entry(
                    aeat_type=aeat_type,
                    classified_at=datetime(2026, 8, 12, 10, 0, 0),
                    classified_by="flask_admin:admin",
                )
                before_snapshot = copy.deepcopy(accounting_entry.invoice.invoice_snapshot)
                before_hash = accounting_entry.invoice.invoice_snapshot_hash
                _, workbook = export_and_open([accounting_entry], tmpdir / "aeat.xlsx")

                self.assertEqual(workbook[AEAT_SALES_LEDGER_SHEET_NAME]["F3"].value, aeat_type)
                self.assertEqual(accounting_entry.invoice.invoice_snapshot, before_snapshot)
                self.assertEqual(accounting_entry.invoice.invoice_snapshot_hash, before_hash)

    def test_legacy_total_rectificative_without_manual_classification_is_rejected_specifically(self):
        accounting_entry = legacy_corrective_entry()
        with temp_export_dir() as tmpdir, self.assertRaisesRegex(
            AeatSalesLedgerValidationError,
            "histórica y requiere clasificación AEAT manual R1/R4",
        ):
            export_aeat_sales_ledger([accounting_entry], output_path=tmpdir / "aeat.xlsx")

    def test_legacy_total_rectificative_with_r4_but_without_audit_is_rejected_specifically(self):
        accounting_entry = legacy_corrective_entry(aeat_type="R4")
        with temp_export_dir() as tmpdir, self.assertRaisesRegex(
            AeatSalesLedgerValidationError,
            "histórica y requiere clasificación AEAT manual R1/R4",
        ):
            export_aeat_sales_ledger([accounting_entry], output_path=tmpdir / "aeat.xlsx")

    def test_legacy_total_rectificative_rejects_missing_audit_invalid_hash_partial_and_out_of_scope_type(self):
        missing_actor = legacy_corrective_entry(
            aeat_type="R1", classified_at=datetime(2026, 8, 12, 10, 0, 0)
        )
        invalid_hash = legacy_corrective_entry(
            aeat_type="R1",
            classified_at=datetime(2026, 8, 12, 10, 0, 0),
            classified_by="flask_admin:admin",
        )
        invalid_hash.invoice.invoice_snapshot_hash = "invalid"
        partial = legacy_corrective_entry(
            aeat_type="R1",
            classified_at=datetime(2026, 8, 12, 10, 0, 0),
            classified_by="flask_admin:admin",
        )
        partial.invoice.invoice_snapshot["operation"]["rectification"]["rectification_scope"] = "partial"
        partial.invoice.invoice_snapshot_hash = calculate_invoice_snapshot_hash(partial.invoice.invoice_snapshot)
        r2 = legacy_corrective_entry(
            aeat_type="R2",
            classified_at=datetime(2026, 8, 12, 10, 0, 0),
            classified_by="flask_admin:admin",
        )

        with temp_export_dir() as tmpdir:
            for label, accounting_entry in (("missing-audit", missing_actor), ("invalid-hash", invalid_hash), ("partial", partial), ("r2", r2)):
                with self.subTest(label=label), self.assertRaises(AeatSalesLedgerValidationError):
                    export_aeat_sales_ledger([accounting_entry], output_path=tmpdir / f"{label}.xlsx")

    def test_mixed_ordinary_and_corrective_rows_keep_deterministic_order(self):
        ordinary = entry(id=1)
        corrective = corrective_entry(id=2)

        with temp_export_dir() as tmpdir:
            _, workbook = export_and_open([corrective, ordinary], tmpdir / "aeat.xlsx")
            sheet = workbook[AEAT_SALES_LEDGER_SHEET_NAME]

        self.assertEqual([sheet["L3"].value, sheet["L4"].value], ["F2026000001", "R2026000001"])

    def test_row_mapping_matches_aeat_sales_ledger_v1(self):
        with temp_export_dir() as tmpdir:
            _, workbook = export_and_open([entry()], tmpdir / "aeat.xlsx")
            sheet = workbook[AEAT_SALES_LEDGER_SHEET_NAME]

            self.assertEqual(sheet["A3"].value, 2026)
            self.assertEqual(sheet["B3"].value, "3T")
            self.assertEqual(sheet["C3"].value, "A")
            self.assertEqual(sheet["D3"].value, "3")
            self.assertEqual(sheet["E3"].value, "3141")
            self.assertEqual(sheet["F3"].value, "F1")
            self.assertEqual(sheet["G3"].value, "I01")
            self.assertEqual(sheet["H3"].value, Decimal("100.00"))
            self.assertEqual(sheet["I3"].value.date(), date(2026, 7, 18))
            self.assertEqual(sheet["J3"].value.date(), date(2026, 7, 17))
            self.assertIsNone(sheet["K3"].value)
            self.assertEqual(sheet["L3"].value, "F2026000001")
            self.assertIsNone(sheet["M3"].value)
            self.assertEqual(sheet["N3"].value, "4")
            self.assertEqual(sheet["O3"].value, "ES")
            self.assertEqual(sheet["P3"].value, "00000000T")
            self.assertEqual(sheet["Q3"].value, "Cliente Fiscal")
            self.assertEqual(sheet["R3"].value, "1")
            self.assertEqual(sheet["S3"].value, "S1")
            self.assertIsNone(sheet["T3"].value)
            self.assertEqual(sheet["U3"].value, Decimal("121.00"))
            self.assertEqual(sheet["V3"].value, Decimal("100.00"))
            self.assertEqual(sheet["W3"].value, Decimal("21.00"))
            self.assertEqual(sheet["X3"].value, Decimal("21.00"))
            self.assertEqual(sheet["AJ3"].value, "order:42")

    def test_dates_amounts_formats_filter_and_freeze_panes_are_configured(self):
        with temp_export_dir() as tmpdir:
            _, workbook = export_and_open([entry()], tmpdir / "aeat.xlsx")
            sheet = workbook[AEAT_SALES_LEDGER_SHEET_NAME]

            self.assertEqual(sheet["I3"].number_format, DATE_FORMAT)
            self.assertEqual(sheet["J3"].number_format, DATE_FORMAT)
            self.assertEqual(sheet["H3"].number_format, MONEY_FORMAT)
            self.assertEqual(sheet["U3"].number_format, MONEY_FORMAT)
            self.assertEqual(sheet["V3"].number_format, MONEY_FORMAT)
            self.assertEqual(sheet["X3"].number_format, MONEY_FORMAT)
            self.assertEqual(sheet.freeze_panes, "A3")
            self.assertEqual(sheet.auto_filter.ref, "A2:AJ3")
            self.assertTrue(all(cell.font.bold for cell in sheet[1]))
            self.assertTrue(all(cell.font.bold for cell in sheet[2]))

    def test_rows_are_sorted_deterministically(self):
        entries = [
            entry(
                id=3,
                invoice_number="F2026000003",
                invoice=invoice(invoice_number="F2026000003"),
            ),
            entry(
                id=2,
                invoice_number="F2026000002",
                invoice=invoice(
                    invoice_number="F2026000002",
                    invoice_snapshot=snapshot(operation={**snapshot()["operation"], "issue_date": "2026-07-17"}),
                ),
            ),
            entry(id=1),
        ]

        with temp_export_dir() as tmpdir:
            _, workbook = export_and_open(entries, tmpdir / "aeat.xlsx")
            sheet = workbook[AEAT_SALES_LEDGER_SHEET_NAME]

            self.assertEqual(
                [sheet.cell(row=row, column=12).value for row in range(3, 6)],
                ["F2026000002", "F2026000001", "F2026000003"],
            )

    def test_generate_workbook_public_helper_does_not_write_file(self):
        workbook = generate_aeat_sales_ledger_workbook([entry()])

        self.assertEqual(workbook.sheetnames, [AEAT_SALES_LEDGER_SHEET_NAME])

    def test_existing_file_without_overwrite_is_rejected(self):
        with temp_export_dir() as tmpdir:
            output_path = tmpdir / "aeat.xlsx"
            output_path.write_bytes(b"existing")

            with self.assertRaises(AeatSalesLedgerWriteError):
                export_aeat_sales_ledger([entry()], output_path=output_path)

            self.assertEqual(output_path.read_bytes(), b"existing")

    def test_overwrite_is_allowed_explicitly(self):
        with temp_export_dir() as tmpdir:
            output_path = tmpdir / "aeat.xlsx"
            output_path.write_bytes(b"existing")

            result = export_aeat_sales_ledger(
                [entry()],
                output_path=output_path,
                overwrite=True,
            )

            self.assertGreater(result.file_size, len(b"existing"))
            self.assertEqual(load_workbook(output_path).sheetnames, [AEAT_SALES_LEDGER_SHEET_NAME])

    def test_invalid_output_path_and_empty_entries_are_rejected(self):
        with temp_export_dir() as tmpdir:
            with self.assertRaises(AeatSalesLedgerValidationError):
                export_aeat_sales_ledger([entry()], output_path=tmpdir / "aeat.xls")
            with self.assertRaises(AeatSalesLedgerValidationError):
                export_aeat_sales_ledger([entry()], output_path="../aeat.xlsx")
            with self.assertRaises(AeatSalesLedgerValidationError):
                export_aeat_sales_ledger([], output_path=tmpdir / "aeat.xlsx")

    def test_legacy_invoice_without_snapshot_is_rejected(self):
        with temp_export_dir() as tmpdir:
            legacy_invoice = InvoiceDouble(
                id=1,
                invoice_number="F2026000001",
                invoice_snapshot=None,
                invoice_snapshot_hash=None,
            )
            legacy_entry = entry(invoice=legacy_invoice)

            with self.assertRaises(AeatSalesLedgerValidationError):
                export_aeat_sales_ledger([legacy_entry], output_path=tmpdir / "aeat.xlsx")

    def test_hash_mismatch_is_rejected(self):
        with temp_export_dir() as tmpdir:
            bad_entry = entry(invoice=invoice(stored_hash="bad-hash"))

            with self.assertRaises(AeatSalesLedgerValidationError):
                export_aeat_sales_ledger([bad_entry], output_path=tmpdir / "aeat.xlsx")

    def test_unsupported_schema_and_legacy_corrective_invoice_are_rejected(self):
        with temp_export_dir() as tmpdir:
            unsupported = snapshot(schema_version=999)
            unsupported_entry = entry(invoice=invoice(invoice_snapshot=unsupported))
            corrective_snapshot = snapshot(
                operation={**snapshot()["operation"], "invoice_type": "corrective"}
            )
            corrective_entry = entry(invoice=invoice(invoice_snapshot=corrective_snapshot))

            with self.assertRaises(AeatSalesLedgerValidationError):
                export_aeat_sales_ledger([unsupported_entry], output_path=tmpdir / "unsupported.xlsx")
            with self.assertRaises(AeatSalesLedgerValidationError):
                export_aeat_sales_ledger([corrective_entry], output_path=tmpdir / "corrective.xlsx")

    def test_corrective_requires_classification_and_matching_frozen_reference(self):
        cases = [
            ("missing model classification", corrective_entry(aeat_type=None)),
            (
                "missing snapshot classification",
                corrective_entry(snapshot_overrides={
                    "operation": {
                        **corrective_snapshot()["operation"],
                        "rectification": {
                            **corrective_snapshot()["operation"]["rectification"],
                            "aeat_type": None,
                        },
                    },
                }),
            ),
            ("r2 out of scope", corrective_entry(aeat_type="R2")),
            ("r3 out of scope", corrective_entry(aeat_type="R3")),
            ("r5 out of scope", corrective_entry(aeat_type="R5")),
            (
                "partial",
                corrective_entry(snapshot_overrides={
                    "operation": {
                        **corrective_snapshot()["operation"],
                        "rectification": {
                            **corrective_snapshot()["operation"]["rectification"],
                            "rectification_scope": "partial",
                        },
                    },
                }),
            ),
        ]

        with temp_export_dir() as tmpdir:
            for label, accounting_entry in cases:
                with self.subTest(label=label), self.assertRaises(AeatSalesLedgerValidationError):
                    export_aeat_sales_ledger([accounting_entry], output_path=tmpdir / f"{label}.xlsx")

    def test_corrective_rejects_model_snapshot_and_original_reference_mismatches(self):
        mismatch = corrective_entry()
        mismatch.invoice.rectification_aeat_type = "R4"
        missing_original = corrective_entry()
        missing_original.invoice.original_invoice = None
        different_original_number = corrective_entry()
        different_original_number.invoice.original_invoice.invoice_number = "F2026000099"
        different_original_date = corrective_entry()
        different_original_date.invoice.original_invoice.issued_at = datetime(2026, 7, 20, 10, 0, 0)
        mismatched_original_id = corrective_entry()
        mismatched_original_id.invoice.original_invoice_id = 99
        v3_ordinary = corrective_entry(snapshot_overrides={
            "operation": {
                **corrective_snapshot()["operation"],
                "invoice_type": "ordinary",
            },
        })
        invalid_hash = corrective_entry()
        invalid_hash.invoice.invoice_snapshot_hash = "invalid"

        with temp_export_dir() as tmpdir:
            for label, accounting_entry in (
                ("type-mismatch", mismatch),
                ("missing-original", missing_original),
                ("number-mismatch", different_original_number),
                ("date-mismatch", different_original_date),
                ("id-mismatch", mismatched_original_id),
                ("v3-ordinary", v3_ordinary),
                ("invalid-hash", invalid_hash),
            ):
                with self.subTest(label=label), self.assertRaises(AeatSalesLedgerValidationError):
                    export_aeat_sales_ledger([accounting_entry], output_path=tmpdir / f"{label}.xlsx")

    def test_non_sale_non_eur_and_missing_tax_id_are_rejected(self):
        with temp_export_dir() as tmpdir:
            with self.assertRaises(AeatSalesLedgerValidationError):
                export_aeat_sales_ledger([entry(entry_type="purchase")], output_path=tmpdir / "type.xlsx")
            with self.assertRaises(AeatSalesLedgerValidationError):
                export_aeat_sales_ledger([entry(currency="USD")], output_path=tmpdir / "currency.xlsx")
            without_tax_id = snapshot(customer={**snapshot()["customer"], "tax_id": None})
            with self.assertRaises(AeatSalesLedgerValidationError):
                export_aeat_sales_ledger(
                    [entry(invoice=invoice(invoice_snapshot=without_tax_id))],
                    output_path=tmpdir / "tax-id.xlsx",
                )

    def test_projection_mismatch_is_rejected(self):
        with temp_export_dir() as tmpdir:
            with self.assertRaises(AeatSalesLedgerValidationError):
                export_aeat_sales_ledger(
                    [entry(total_amount=Decimal("120.99"))],
                    output_path=tmpdir / "amount.xlsx",
                )
            with self.assertRaises(AeatSalesLedgerValidationError):
                export_aeat_sales_ledger(
                    [entry(invoice_number="F2026000099")],
                    output_path=tmpdir / "number.xlsx",
                )

    def test_multiple_vat_rates_are_rejected(self):
        with temp_export_dir() as tmpdir:
            multi_rate = snapshot(
                lines=[
                    {**snapshot()["lines"][0], "tax_rate": "21.00"},
                    {**snapshot()["lines"][0], "line_number": 2, "tax_rate": "10.00"},
                ]
            )

            with self.assertRaises(AeatSalesLedgerValidationError):
                export_aeat_sales_ledger(
                    [entry(invoice=invoice(invoice_snapshot=multi_rate))],
                    output_path=tmpdir / "aeat.xlsx",
                )

    def test_operation_date_falls_back_to_snapshot_issue_date(self):
        operation = copy.deepcopy(snapshot()["operation"])
        operation.pop("operation_date")
        fallback_snapshot = snapshot(operation=operation)

        with temp_export_dir() as tmpdir:
            _, workbook = export_and_open(
                [entry(invoice=invoice(invoice_snapshot=fallback_snapshot))],
                tmpdir / "aeat.xlsx",
            )
            sheet = workbook[AEAT_SALES_LEDGER_SHEET_NAME]

            self.assertEqual(sheet["J3"].value.date(), date(2026, 7, 18))

    def test_export_does_not_mutate_entries_or_invoice_snapshot(self):
        invoice_double = invoice()
        before_snapshot = copy.deepcopy(invoice_double.invoice_snapshot)
        accounting_entry = entry(invoice=invoice_double)

        with temp_export_dir() as tmpdir:
            export_aeat_sales_ledger([accounting_entry], output_path=tmpdir / "aeat.xlsx")

        self.assertEqual(invoice_double.invoice_snapshot, before_snapshot)
        self.assertEqual(accounting_entry.status, "pending")


if __name__ == "__main__":
    unittest.main()
