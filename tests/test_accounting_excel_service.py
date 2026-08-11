import shutil
import sys
import unittest
import uuid
from contextlib import contextmanager
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
SRC_DIR = ROOT_DIR / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))


from api.accounting_excel_service import (  # noqa: E402
    DATE_FORMAT,
    HEADERS,
    MONEY_FORMAT,
    SALES_SHEET_NAME,
    AccountingExcelValidationError,
    AccountingExcelWriteError,
    export_sales_accounting_entries,
)


TEST_TMP_ROOT = ROOT_DIR / "tmp-accounting-excel-tests"


@dataclass
class InvoiceDouble:
    invoice_type: str | None = "ordinary"
    original_invoice_id: int | None = None
    original_invoice: object | None = None

    @property
    def order(self):
        raise AssertionError("Excel export must not read live order relationship")

    @property
    def invoice_snapshot(self):
        raise AssertionError("Excel export must not read invoice snapshot")


@dataclass
class AccountingEntryDouble:
    id: int
    invoice_date: date
    invoice_number: str
    customer_name: str = "Cliente Contable"
    customer_tax_id: str | None = "00000000T"
    taxable_base: Decimal = Decimal("100.00")
    vat_amount: Decimal = Decimal("21.00")
    total_amount: Decimal = Decimal("121.00")
    currency: str = "EUR"
    payment_provider: str | None = "stripe"
    order_id: int | None = 123
    status: str = "pending"
    entry_type: str = "sale"

    invoice: InvoiceDouble = field(default_factory=InvoiceDouble)

    @property
    def order(self):
        raise AssertionError("Excel export must not read live order relationship")

    @property
    def user(self):
        raise AssertionError("Excel export must not read live user relationship")

    @property
    def checkout_session(self):
        raise AssertionError("Excel export must not read live checkout relationship")

    @property
    def invoice_snapshot(self):
        raise AssertionError("Excel export must not read invoice snapshot")


@contextmanager
def temp_export_dir():
    TEST_TMP_ROOT.mkdir(exist_ok=True)
    case_dir = TEST_TMP_ROOT / f"case-{uuid.uuid4().hex}"
    case_dir.mkdir()
    try:
        yield case_dir
    finally:
        shutil.rmtree(case_dir, ignore_errors=True)


def tearDownModule():
    shutil.rmtree(TEST_TMP_ROOT, ignore_errors=True)


def entry(**overrides):
    data = {
        "id": 1,
        "invoice_date": date(2026, 7, 15),
        "invoice_number": "F2026000001",
    }
    data.update(overrides)
    return AccountingEntryDouble(**data)


def export_and_open(entries, path):
    result = export_sales_accounting_entries(entries, output_path=path)
    workbook = load_workbook(path, data_only=True)
    return result, workbook


class AccountingExcelServiceTest(unittest.TestCase):
    def test_valid_export_creates_readable_xlsx(self):
        with temp_export_dir() as tmpdir:
            output_path = tmpdir / "ingresos.xlsx"
            result, workbook = export_and_open([entry()], output_path)

            self.assertTrue(output_path.exists())
            self.assertEqual(result.filename, "ingresos.xlsx")
            self.assertEqual(result.row_count, 1)
            self.assertGreater(result.file_size, 0)
            self.assertIn(SALES_SHEET_NAME, workbook.sheetnames)

    def test_sheet_name_and_exact_headers(self):
        with temp_export_dir() as tmpdir:
            _, workbook = export_and_open([entry()], tmpdir / "ingresos.xlsx")
            sheet = workbook[SALES_SHEET_NAME]

            self.assertEqual([cell.value for cell in sheet[1]], HEADERS)

    def test_rows_are_sorted_deterministically(self):
        entries = [
            entry(id=3, invoice_date=date(2026, 7, 16), invoice_number="F2026000003"),
            entry(id=2, invoice_date=date(2026, 7, 15), invoice_number="F2026000002"),
            entry(id=1, invoice_date=date(2026, 7, 15), invoice_number="F2026000001"),
        ]

        with temp_export_dir() as tmpdir:
            _, workbook = export_and_open(entries, tmpdir / "ingresos.xlsx")
            sheet = workbook[SALES_SHEET_NAME]

            self.assertEqual(
                [sheet.cell(row=row, column=2).value for row in range(2, 5)],
                ["F2026000001", "F2026000002", "F2026000003"],
            )

    def test_amount_values_are_numeric_and_correct(self):
        with temp_export_dir() as tmpdir:
            _, workbook = export_and_open([entry()], tmpdir / "ingresos.xlsx")
            sheet = workbook[SALES_SHEET_NAME]

            self.assertEqual(sheet["G2"].value, Decimal("100.00"))
            self.assertEqual(sheet["H2"].value, Decimal("21.00"))
            self.assertEqual(sheet["I2"].value, Decimal("121.00"))

    def test_dates_are_written_as_dates(self):
        with temp_export_dir() as tmpdir:
            _, workbook = export_and_open([entry()], tmpdir / "ingresos.xlsx")
            sheet = workbook[SALES_SHEET_NAME]

            self.assertEqual(sheet["A2"].value.date(), date(2026, 7, 15))

    def test_filter_and_freeze_panes_are_configured(self):
        with temp_export_dir() as tmpdir:
            _, workbook = export_and_open([entry()], tmpdir / "ingresos.xlsx")
            sheet = workbook[SALES_SHEET_NAME]

            self.assertEqual(sheet.freeze_panes, "A2")
            self.assertEqual(sheet.auto_filter.ref, "A1:M2")

    def test_money_and_date_formats_are_applied(self):
        with temp_export_dir() as tmpdir:
            _, workbook = export_and_open([entry()], tmpdir / "ingresos.xlsx")
            sheet = workbook[SALES_SHEET_NAME]

            self.assertEqual(sheet["A2"].number_format, DATE_FORMAT)
            self.assertEqual(sheet["G2"].number_format, MONEY_FORMAT)
            self.assertEqual(sheet["H2"].number_format, MONEY_FORMAT)
            self.assertEqual(sheet["I2"].number_format, MONEY_FORMAT)

    def test_header_bold_and_column_widths_are_set(self):
        with temp_export_dir() as tmpdir:
            _, workbook = export_and_open([entry()], tmpdir / "ingresos.xlsx")
            sheet = workbook[SALES_SHEET_NAME]

            self.assertTrue(all(cell.font.bold for cell in sheet[1]))
            self.assertGreaterEqual(sheet.column_dimensions["B"].width, 20)

    def test_invoice_without_tax_id_exports_empty_cell(self):
        with temp_export_dir() as tmpdir:
            _, workbook = export_and_open([entry(customer_tax_id=None)], tmpdir / "ingresos.xlsx")
            sheet = workbook[SALES_SHEET_NAME]

            self.assertIsNone(sheet["F2"].value)

    def test_stripe_payment_provider_is_exported(self):
        with temp_export_dir() as tmpdir:
            _, workbook = export_and_open([entry(payment_provider="stripe")], tmpdir / "ingresos.xlsx")
            sheet = workbook[SALES_SHEET_NAME]

            self.assertEqual(sheet["K2"].value, "stripe")

    def test_paypal_payment_provider_is_exported(self):
        with temp_export_dir() as tmpdir:
            _, workbook = export_and_open([entry(payment_provider="paypal")], tmpdir / "ingresos.xlsx")
            sheet = workbook[SALES_SHEET_NAME]

            self.assertEqual(sheet["K2"].value, "paypal")

    def test_nullable_order_id_exports_empty_cell(self):
        with temp_export_dir() as tmpdir:
            _, workbook = export_and_open([entry(order_id=None)], tmpdir / "ingresos.xlsx")
            sheet = workbook[SALES_SHEET_NAME]

            self.assertIsNone(sheet["L2"].value)

    def test_ordinary_invoice_is_identified_without_rectification_reference(self):
        with temp_export_dir() as tmpdir:
            _, workbook = export_and_open([entry()], tmpdir / "ingresos.xlsx")
            sheet = workbook[SALES_SHEET_NAME]

            self.assertEqual(sheet["C2"].value, "Ordinaria")
            self.assertIsNone(sheet["D2"].value)

    def test_corrective_invoice_keeps_reference_and_negative_amounts(self):
        corrective = InvoiceDouble(
            invoice_type="corrective",
            original_invoice_id=50,
            original_invoice=type("OriginalInvoice", (), {"invoice_number": "F2026000001"})(),
        )
        corrective_entry = entry(
            invoice_number="R2026000001",
            taxable_base=Decimal("-100.00"),
            vat_amount=Decimal("-21.00"),
            total_amount=Decimal("-121.00"),
            invoice=corrective,
        )

        with temp_export_dir() as tmpdir:
            _, workbook = export_and_open([corrective_entry], tmpdir / "ingresos.xlsx")
            sheet = workbook[SALES_SHEET_NAME]

            self.assertEqual(sheet["C2"].value, "Rectificativa")
            self.assertEqual(sheet["D2"].value, "F2026000001")
            self.assertEqual(sheet["G2"].value, Decimal("-100.00"))
            self.assertEqual(sheet["H2"].value, Decimal("-21.00"))
            self.assertEqual(sheet["I2"].value, Decimal("-121.00"))

    def test_existing_file_without_overwrite_is_rejected(self):
        with temp_export_dir() as tmpdir:
            output_path = tmpdir / "ingresos.xlsx"
            output_path.write_bytes(b"existing")

            with self.assertRaises(AccountingExcelWriteError):
                export_sales_accounting_entries([entry()], output_path=output_path)

            self.assertEqual(output_path.read_bytes(), b"existing")

    def test_overwrite_is_allowed_explicitly(self):
        with temp_export_dir() as tmpdir:
            output_path = tmpdir / "ingresos.xlsx"
            output_path.write_bytes(b"existing")

            result = export_sales_accounting_entries(
                [entry()],
                output_path=output_path,
                overwrite=True,
            )

            self.assertGreater(result.file_size, len(b"existing"))
            self.assertEqual(load_workbook(output_path).sheetnames, [SALES_SHEET_NAME])

    def test_invalid_extension_is_rejected(self):
        with temp_export_dir() as tmpdir:
            with self.assertRaises(AccountingExcelValidationError):
                export_sales_accounting_entries([entry()], output_path=tmpdir / "ingresos.xls")

    def test_empty_output_path_is_rejected(self):
        with self.assertRaises(AccountingExcelValidationError):
            export_sales_accounting_entries([entry()], output_path="")

    def test_path_traversal_is_rejected(self):
        with self.assertRaises(AccountingExcelValidationError):
            export_sales_accounting_entries([entry()], output_path="../ingresos.xlsx")

    def test_missing_entries_are_rejected(self):
        with temp_export_dir() as tmpdir:
            with self.assertRaises(AccountingExcelValidationError):
                export_sales_accounting_entries([], output_path=tmpdir / "ingresos.xlsx")

    def test_missing_invoice_number_is_rejected(self):
        with temp_export_dir() as tmpdir:
            with self.assertRaises(AccountingExcelValidationError):
                export_sales_accounting_entries(
                    [entry(invoice_number="")],
                    output_path=tmpdir / "ingresos.xlsx",
                )

    def test_invalid_amount_is_rejected(self):
        with temp_export_dir() as tmpdir:
            with self.assertRaises(AccountingExcelValidationError):
                export_sales_accounting_entries(
                    [entry(taxable_base="not-money")],
                    output_path=tmpdir / "ingresos.xlsx",
                )

    def test_non_sale_entry_type_is_rejected(self):
        with temp_export_dir() as tmpdir:
            with self.assertRaises(AccountingExcelValidationError):
                export_sales_accounting_entries(
                    [entry(entry_type="purchase")],
                    output_path=tmpdir / "ingresos.xlsx",
                )

    def test_non_eur_currency_is_rejected(self):
        with temp_export_dir() as tmpdir:
            with self.assertRaises(AccountingExcelValidationError):
                export_sales_accounting_entries(
                    [entry(currency="USD")],
                    output_path=tmpdir / "ingresos.xlsx",
                )

    def test_export_reads_only_the_persisted_invoice_relation(self):
        with temp_export_dir() as tmpdir:
            export_sales_accounting_entries([entry()], output_path=tmpdir / "ingresos.xlsx")

    def test_export_does_not_mutate_entries_or_status(self):
        accounting_entry = entry(status="pending")
        before = accounting_entry.__dict__.copy()

        with temp_export_dir() as tmpdir:
            export_sales_accounting_entries([accounting_entry], output_path=tmpdir / "ingresos.xlsx")

        self.assertEqual(accounting_entry.__dict__, before)
        self.assertEqual(accounting_entry.status, "pending")

    def test_service_does_not_commit_or_query_live_models(self):
        source = (SRC_DIR / "api/accounting_excel_service.py").read_text(encoding="utf-8")

        for forbidden in (
            "db.session",
            ".commit(",
            ".rollback(",
            "InvoiceSnapshot",
            "OrderDetails",
            "Orders",
            "Users",
            "CheckoutSessions",
            "Products",
            "entry.order",
            "invoice.order",
            "invoice.invoice_snapshot",
        ):
            self.assertNotIn(forbidden, source)

    def test_service_has_no_external_effects_or_master_excel_logic(self):
        source = (SRC_DIR / "api/accounting_excel_service.py").read_text(encoding="utf-8").lower()

        for forbidden in (
            "requests",
            "pandas",
            "libreoffice",
            "verifactu",
            "ocr",
            "modelo 303",
            "modelo 390",
            "send_email",
        ):
            self.assertNotIn(forbidden, source)

    def test_write_error_is_sanitized(self):
        with temp_export_dir() as tmpdir:
            file_parent = tmpdir / "not-a-directory"
            file_parent.write_text("x", encoding="utf-8")

            with self.assertRaisesRegex(AccountingExcelWriteError, "No se pudo escribir"):
                export_sales_accounting_entries(
                    [entry()],
                    output_path=file_parent / "ingresos.xlsx",
                )


if __name__ == "__main__":
    unittest.main()
