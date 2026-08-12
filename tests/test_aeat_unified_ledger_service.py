import copy
import sys
import unittest
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace
from tempfile import TemporaryDirectory


ROOT_DIR = Path(__file__).resolve().parents[1]
SRC_DIR = ROOT_DIR / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))


from api.aeat_received_expense_ledger_xlsx_service import (  # noqa: E402
    generate_aeat_received_expense_ledger_workbook,
)
from api.aeat_sales_ledger_service import generate_aeat_sales_ledger_workbook  # noqa: E402
from api.aeat_unified_ledger_service import (  # noqa: E402
    AeatUnifiedLedgerValidationError,
    export_aeat_unified_ledger,
    generate_aeat_unified_ledger_workbook,
)
from api.invoice_snapshot_integrity import calculate_invoice_snapshot_hash  # noqa: E402
from api.supplier_invoice_snapshot_integrity import (  # noqa: E402
    calculate_supplier_invoice_snapshot_hash,
)


@dataclass
class InvoiceDouble:
    id: int
    invoice_number: str
    invoice_snapshot: dict
    invoice_snapshot_hash: str
    issued_at: datetime
    invoice_type: str = "ordinary"
    original_invoice_id: int | None = None
    rectification_aeat_type: str | None = None
    rectification_aeat_classified_at: datetime | None = None
    rectification_aeat_classified_by: str | None = None
    original_invoice: object | None = None


@dataclass
class AccountingEntryDouble:
    id: int
    invoice_date: date
    invoice_number: str
    invoice: InvoiceDouble
    entry_type: str = "sale"
    currency: str = "EUR"
    taxable_base: Decimal = Decimal("100.00")
    vat_amount: Decimal = Decimal("21.00")
    total_amount: Decimal = Decimal("121.00")


def sales_snapshot(issue_date="2026-02-10"):
    return {
        "schema_version": 2,
        "issuer": {"legal_name": "MetalWolft", "tax_id": "B00000000"},
        "customer": {
            "legal_name": "Cliente Fiscal",
            "tax_id": "00000000T",
            "country_code": "ES",
        },
        "operation": {
            "invoice_type": "ordinary",
            "issue_date": issue_date,
            "operation_date": issue_date,
            "currency": "EUR",
            "order_id": 42,
        },
        "lines": [{
            "line_number": 1,
            "tax_rate": "21.00",
            "tax_base": "100.00",
            "tax_amount": "21.00",
            "line_total": "121.00",
        }],
        "totals": {"tax_base": "100.00", "tax_amount": "21.00", "total_amount": "121.00"},
        "payment": {"provider": "stripe"},
        "references": {"order_id": 42},
    }


def sales_entry(entry_id=1, issue_date="2026-02-10"):
    snapshot = sales_snapshot(issue_date)
    invoice_number = f"F2026{entry_id:06d}"
    invoice = InvoiceDouble(
        id=entry_id,
        invoice_number=invoice_number,
        invoice_snapshot=snapshot,
        invoice_snapshot_hash=calculate_invoice_snapshot_hash(snapshot),
        issued_at=datetime.fromisoformat(f"{issue_date}T10:00:00"),
    )
    return AccountingEntryDouble(
        id=entry_id,
        invoice_date=date.fromisoformat(issue_date),
        invoice_number=invoice_number,
        invoice=invoice,
    )


def legacy_corrective_entry():
    original_entry = sales_entry(1, "2026-02-10")
    original = original_entry.invoice
    snapshot = copy.deepcopy(original.invoice_snapshot)
    snapshot["schema_version"] = 3
    snapshot["operation"] = {
        **snapshot["operation"],
        "invoice_type": "corrective",
        "issue_date": "2026-05-10",
        "operation_date": "2026-05-10",
        "rectification": {
            "rectification_type": "differences",
            "rectification_scope": "total",
            "rectification_reason": "invoice_error",
            "rectification_reason_text": "Factura emitida por error",
            "original_invoice_id": original.id,
            "original_invoice_number": original.invoice_number,
            "original_invoice_issued_at": original.issued_at.isoformat(),
            "affected_line_numbers": [1],
        },
    }
    snapshot["lines"][0].update(tax_base="-100.00", tax_amount="-21.00", line_total="-121.00")
    snapshot["totals"] = {"tax_base": "-100.00", "tax_amount": "-21.00", "total_amount": "-121.00"}
    invoice = InvoiceDouble(
        id=2,
        invoice_number="R2026000001",
        invoice_snapshot=snapshot,
        invoice_snapshot_hash=calculate_invoice_snapshot_hash(snapshot),
        issued_at=datetime(2026, 5, 10, 10, 0, 0),
        invoice_type="corrective",
        original_invoice_id=original.id,
        rectification_aeat_type="R4",
        rectification_aeat_classified_at=datetime(2026, 8, 12, 10, 0, 0),
        rectification_aeat_classified_by="flask_admin:admin",
        original_invoice=original,
    )
    return AccountingEntryDouble(
        id=2,
        invoice_date=date(2026, 5, 10),
        invoice_number=invoice.invoice_number,
        invoice=invoice,
        taxable_base=Decimal("-100.00"),
        vat_amount=Decimal("-21.00"),
        total_amount=Decimal("-121.00"),
    )


def supplier_invoice(reception_number=1, received_at="2026-02-12T00:00:00"):
    snapshot = {
        "schema_version": 2,
        "supplier": {
            "legal_name": "Proveedor Nacional S.L.",
            "tax_id": "B13019559",
            "country_code": "ES",
            "tax_id_type": "NIF",
        },
        "document": {
            "supplier_invoice_number": f"PROV-{reception_number}",
            "reception_number": reception_number,
            "issue_date": received_at[:10],
            "operation_date": received_at[:10],
            "received_at": received_at,
            "currency": "EUR",
            "fiscal_invoice_type": "F1",
            "tax_treatment": "domestic_standard",
            "special_regime_key": None,
        },
        "tax_breakdowns": [{
            "position": 1,
            "tax_base": "35.76",
            "tax_rate": "21.00",
            "tax_amount": "7.51",
            "deductible_tax_amount": "7.51",
        }],
        "totals": {
            "tax_base": "35.76",
            "tax_amount": "7.51",
            "deductible_tax_amount": "7.51",
            "total_amount": "43.27",
        },
        "expense_classification": {
            "aeat_expense_concept_code": "G01",
            "expense_deductible_amount": "35.76",
        },
    }
    return SimpleNamespace(
        status="registered",
        snapshot_schema_version=2,
        fiscal_snapshot=snapshot,
        snapshot_hash=calculate_supplier_invoice_snapshot_hash(snapshot),
        reception_number=reception_number,
    )


def sheet_values(worksheet):
    return [[cell.value for cell in row] for row in worksheet.iter_rows()]


class AeatUnifiedLedgerServiceTest(unittest.TestCase):
    def test_combined_workbook_preserves_sheet_order_and_independent_content(self):
        sale = sales_entry()
        received = supplier_invoice()
        workbook, sales_rows, received_rows = generate_aeat_unified_ledger_workbook(
            [sale], [received], year=2026, period="1T"
        )
        sales_workbook = generate_aeat_sales_ledger_workbook([sale])
        received_workbook, _ = generate_aeat_received_expense_ledger_workbook(
            [received], year=2026, period="1T"
        )

        self.assertEqual(workbook.sheetnames, ["EXPEDIDAS_INGRESOS", "RECIBIDAS_GASTOS"])
        self.assertEqual(sheet_values(workbook["EXPEDIDAS_INGRESOS"]), sheet_values(sales_workbook.active))
        self.assertEqual(sheet_values(workbook["RECIBIDAS_GASTOS"]), sheet_values(received_workbook.active))
        self.assertEqual(len(sales_rows), 1)
        self.assertEqual(len(received_rows), 1)

    def test_one_empty_domain_still_creates_both_sheets(self):
        sales_only, _, received_rows = generate_aeat_unified_ledger_workbook(
            [sales_entry()], [], year=2026, period="1T"
        )
        received_only, sales_rows, _ = generate_aeat_unified_ledger_workbook(
            [], [supplier_invoice()], year=2026, period="1T"
        )

        self.assertEqual(sales_only.sheetnames, ["EXPEDIDAS_INGRESOS", "RECIBIDAS_GASTOS"])
        self.assertEqual(sales_only["RECIBIDAS_GASTOS"].max_row, 2)
        self.assertEqual(received_rows, [])
        self.assertEqual(received_only["EXPEDIDAS_INGRESOS"].max_row, 2)
        self.assertEqual(sales_rows, [])

    def test_both_empty_or_outside_requested_period_is_rejected(self):
        with self.assertRaisesRegex(AeatUnifiedLedgerValidationError, "No hay operaciones"):
            generate_aeat_unified_ledger_workbook([], [], year=2026, period="1T")
        with self.assertRaisesRegex(AeatUnifiedLedgerValidationError, "No hay operaciones"):
            generate_aeat_unified_ledger_workbook(
                [sales_entry(issue_date="2025-02-10")],
                [supplier_invoice(received_at="2025-02-12T00:00:00")],
                year=2026,
                period="4T",
            )

    def test_period_selection_is_cumulative_for_both_domains(self):
        sales = [
            sales_entry(quarter, issue_date)
            for quarter, issue_date in enumerate(
                ("2026-02-01", "2026-05-01", "2026-08-01", "2026-11-01"), start=1
            )
        ]
        received = [
            supplier_invoice(quarter, f"2026-{month:02d}-01T00:00:00")
            for quarter, month in enumerate((2, 5, 8, 11), start=1)
        ]
        for expected_count, period in enumerate(("1T", "2T", "3T", "4T"), start=1):
            with self.subTest(period=period):
                _, sales_rows, received_rows = generate_aeat_unified_ledger_workbook(
                    sales, received, year=2026, period=period
                )
                self.assertEqual(len(sales_rows), expected_count)
                self.assertEqual(len(received_rows), expected_count)
                self.assertEqual([row["period"] for row in sales_rows], [f"{value}T" for value in range(1, expected_count + 1)])

    def test_classified_legacy_rectification_is_included_and_missing_audit_is_rejected(self):
        legacy = legacy_corrective_entry()
        workbook, sales_rows, _ = generate_aeat_unified_ledger_workbook(
            [legacy], [], year=2026, period="2T"
        )
        self.assertEqual(workbook["EXPEDIDAS_INGRESOS"]["F3"].value, "R4")
        self.assertEqual(len(sales_rows), 1)

        legacy.invoice.rectification_aeat_classified_at = None
        legacy.invoice.rectification_aeat_classified_by = None
        with self.assertRaisesRegex(AeatUnifiedLedgerValidationError, "requiere clasificación AEAT manual"):
            generate_aeat_unified_ledger_workbook([legacy], [], year=2026, period="2T")

    def test_invalid_snapshot_in_either_domain_aborts_the_whole_workbook(self):
        invalid_sale = sales_entry()
        invalid_sale.invoice.invoice_snapshot_hash = "invalid"
        with self.assertRaisesRegex(AeatUnifiedLedgerValidationError, "integridad"):
            generate_aeat_unified_ledger_workbook(
                [invalid_sale], [supplier_invoice()], year=2026, period="1T"
            )

        invalid_received = supplier_invoice()
        invalid_received.snapshot_hash = "invalid"
        with self.assertRaisesRegex(AeatUnifiedLedgerValidationError, "integridad"):
            generate_aeat_unified_ledger_workbook(
                [sales_entry()], [invalid_received], year=2026, period="1T"
            )

    def test_unclassified_received_v1_aborts_the_workbook(self):
        legacy_received = supplier_invoice()
        legacy_received.fiscal_snapshot.pop("expense_classification")
        legacy_received.fiscal_snapshot["document"].pop("received_at")
        legacy_received.fiscal_snapshot["schema_version"] = 1
        legacy_received.snapshot_schema_version = 1
        legacy_received.snapshot_hash = calculate_supplier_invoice_snapshot_hash(
            legacy_received.fiscal_snapshot
        )

        with self.assertRaisesRegex(AeatUnifiedLedgerValidationError, "requiere clasificaci"):
            generate_aeat_unified_ledger_workbook(
                [sales_entry()], [legacy_received], year=2026, period="1T"
            )

    def test_classified_received_v1_is_included_without_changing_the_snapshot(self):
        legacy_received = supplier_invoice()
        legacy_received.fiscal_snapshot.pop("expense_classification")
        legacy_received.fiscal_snapshot["document"].pop("received_at")
        legacy_received.fiscal_snapshot["schema_version"] = 1
        legacy_received.snapshot_schema_version = 1
        legacy_received.snapshot_hash = calculate_supplier_invoice_snapshot_hash(
            legacy_received.fiscal_snapshot
        )
        legacy_received.aeat_expense_concept_code = "G01"
        legacy_received.expense_deductible_amount = Decimal("35.76")
        legacy_received.legacy_expense_received_at = date(2026, 2, 12)
        legacy_received.legacy_expense_classified_at = datetime(2026, 8, 12, 10, 0, 0)
        legacy_received.legacy_expense_classified_by = "flask_admin:admin"
        snapshot_before = copy.deepcopy(legacy_received.fiscal_snapshot)
        hash_before = legacy_received.snapshot_hash

        workbook, _, received_rows = generate_aeat_unified_ledger_workbook(
            [sales_entry()], [legacy_received], year=2026, period="1T"
        )

        self.assertEqual(workbook["RECIBIDAS_GASTOS"]["G3"].value, "G01")
        self.assertEqual(len(received_rows), 1)
        self.assertEqual(legacy_received.fiscal_snapshot, snapshot_before)
        self.assertEqual(legacy_received.snapshot_hash, hash_before)

    def test_export_writes_the_requested_two_sheet_xlsx(self):
        with TemporaryDirectory() as temporary_dir:
            output_path = Path(temporary_dir) / "metalwolft_aeat_2026_1T.xlsx"
            result = export_aeat_unified_ledger(
                [sales_entry()],
                [supplier_invoice()],
                year=2026,
                period="1T",
                output_path=output_path,
            )
            from openpyxl import load_workbook

            workbook = load_workbook(output_path, data_only=True)

        self.assertEqual(result.filename, "metalwolft_aeat_2026_1T.xlsx")
        self.assertEqual(result.sales_row_count, 1)
        self.assertEqual(result.received_invoice_count, 1)
        self.assertEqual(result.received_row_count, 1)
        self.assertEqual(workbook.sheetnames, ["EXPEDIDAS_INGRESOS", "RECIBIDAS_GASTOS"])


if __name__ == "__main__":
    unittest.main()
