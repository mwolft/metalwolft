from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path


SALES_SHEET_NAME = "Ingresos"
SALE_ENTRY_TYPE = "sale"
SUPPORTED_CURRENCY = "EUR"
HEADERS = [
    "Fecha factura",
    "Número factura",
    "Tipo de factura",
    "Factura rectificada",
    "Cliente",
    "NIF/CIF",
    "Base imponible",
    "IVA",
    "Total",
    "Moneda",
    "Método de pago",
    "Pedido",
    "Estado contable",
]
DATE_FORMAT = "DD/MM/YYYY"
MONEY_FORMAT = '#,##0.00'


class AccountingExcelExportError(Exception):
    """Base error for accounting Excel exports."""


class AccountingExcelValidationError(AccountingExcelExportError):
    """Raised when export input is not valid."""


class AccountingExcelWriteError(AccountingExcelExportError):
    """Raised when the export file cannot be written safely."""


@dataclass(frozen=True)
class AccountingExcelExportResult:
    output_path: str
    filename: str
    row_count: int
    generated_at: datetime
    file_size: int


def export_sales_accounting_entries(entries, *, output_path, overwrite=False):
    """Export sale AccountingEntry records to a deterministic XLSX workbook."""
    generated_at = datetime.now(timezone.utc)
    path = _validated_output_path(output_path)
    prepared_entries = _prepared_entries(entries)

    if path.exists() and not overwrite:
        raise AccountingExcelWriteError("El archivo de exportacion contable ya existe.")

    workbook = _build_workbook(prepared_entries)

    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
    except OSError as exc:
        raise AccountingExcelWriteError("No se pudo escribir el Excel de ingresos.") from exc

    return AccountingExcelExportResult(
        output_path=str(path),
        filename=path.name,
        row_count=len(prepared_entries),
        generated_at=generated_at,
        file_size=path.stat().st_size,
    )


def _validated_output_path(output_path):
    if not output_path:
        raise AccountingExcelValidationError("La ruta de salida es obligatoria.")

    path = Path(output_path)
    if any(part == ".." for part in path.parts):
        raise AccountingExcelValidationError("La ruta de salida no es valida.")
    if path.suffix.lower() != ".xlsx":
        raise AccountingExcelValidationError("La exportacion debe tener extension .xlsx.")
    if not path.name or path.name == ".xlsx":
        raise AccountingExcelValidationError("El nombre del archivo de salida no es valido.")
    return path


def _prepared_entries(entries):
    if entries is None:
        raise AccountingExcelValidationError("Debe indicarse al menos un registro contable.")

    prepared = [_prepared_entry(entry) for entry in list(entries)]
    if not prepared:
        raise AccountingExcelValidationError("Debe indicarse al menos un registro contable.")

    return sorted(
        prepared,
        key=lambda entry: (
            entry["invoice_date"],
            entry["invoice_number"],
            entry["id"],
        ),
    )


def _prepared_entry(entry):
    entry_type = _required_text(getattr(entry, "entry_type", None), "entry_type")
    if entry_type != SALE_ENTRY_TYPE:
        raise AccountingExcelValidationError("Solo se pueden exportar registros contables de venta.")

    currency = _required_text(getattr(entry, "currency", None), "currency").upper()
    if currency != SUPPORTED_CURRENCY:
        raise AccountingExcelValidationError("Moneda no soportada para la exportacion contable.")

    document = _document_context(entry)

    return {
        "id": _entry_id(entry),
        "invoice_date": _invoice_date(getattr(entry, "invoice_date", None)),
        "invoice_number": _required_text(getattr(entry, "invoice_number", None), "invoice_number"),
        "invoice_type": document["invoice_type"],
        "rectified_invoice_number": document["rectified_invoice_number"],
        "customer_name": _required_text(getattr(entry, "customer_name", None), "customer_name"),
        "customer_tax_id": _optional_text(getattr(entry, "customer_tax_id", None)),
        "taxable_base": _money(getattr(entry, "taxable_base", None), "taxable_base"),
        "vat_amount": _money(getattr(entry, "vat_amount", None), "vat_amount"),
        "total_amount": _money(getattr(entry, "total_amount", None), "total_amount"),
        "currency": currency,
        "payment_provider": _optional_text(getattr(entry, "payment_provider", None)),
        "order_id": getattr(entry, "order_id", None),
        "status": _required_text(getattr(entry, "status", None), "status"),
    }


def _build_workbook(entries):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SALES_SHEET_NAME
    worksheet.append(HEADERS)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for entry in entries:
        worksheet.append([
            entry["invoice_date"],
            entry["invoice_number"],
            entry["invoice_type"],
            entry["rectified_invoice_number"],
            entry["customer_name"],
            entry["customer_tax_id"],
            entry["taxable_base"],
            entry["vat_amount"],
            entry["total_amount"],
            entry["currency"],
            entry["payment_provider"],
            entry["order_id"],
            entry["status"],
        ])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _apply_formats(worksheet)
    _apply_column_widths(worksheet)
    return workbook


def _apply_formats(worksheet):
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        row[0].number_format = DATE_FORMAT
        for cell in row[6:9]:
            cell.number_format = MONEY_FORMAT


def _apply_column_widths(worksheet):
    widths = {
        "A": 14,
        "B": 22,
        "C": 18,
        "D": 22,
        "E": 28,
        "F": 16,
        "G": 16,
        "H": 12,
        "I": 14,
        "J": 10,
        "K": 18,
        "L": 12,
        "M": 18,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def _entry_id(entry):
    entry_id = getattr(entry, "id", None)
    return int(entry_id or 0)


def _document_context(entry):
    invoice = getattr(entry, "invoice", None)
    if invoice is None:
        raise AccountingExcelValidationError("La factura emitida es obligatoria para la exportacion.")

    invoice_type = getattr(invoice, "invoice_type", None)
    if invoice_type in (None, "ordinary"):
        return {
            "invoice_type": "Ordinaria",
            "rectified_invoice_number": None,
        }
    if invoice_type != "corrective":
        raise AccountingExcelValidationError("Tipo de factura no soportado para la exportacion.")

    original_invoice_id = getattr(invoice, "original_invoice_id", None)
    original_invoice = getattr(invoice, "original_invoice", None)
    if not original_invoice_id or original_invoice is None:
        raise AccountingExcelValidationError("La factura rectificativa no tiene factura original valida.")

    return {
        "invoice_type": "Rectificativa",
        "rectified_invoice_number": _required_text(
            getattr(original_invoice, "invoice_number", None),
            "original_invoice.invoice_number",
        ),
    }


def _invoice_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value:
        try:
            return datetime.fromisoformat(value).date()
        except ValueError as exc:
            raise AccountingExcelValidationError("Fecha de factura invalida.") from exc
    raise AccountingExcelValidationError("Fecha de factura obligatoria.")


def _money(value, field):
    try:
        amount = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise AccountingExcelValidationError(f"Importe no valido en {field}.") from exc
    if not amount.is_finite():
        raise AccountingExcelValidationError(f"Importe no valido en {field}.")
    return amount.quantize(Decimal("0.01"))


def _required_text(value, field):
    text = _optional_text(value)
    if not text:
        raise AccountingExcelValidationError(f"Campo obligatorio ausente: {field}.")
    return text


def _optional_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None
