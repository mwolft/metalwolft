"""XLSX writer for the isolated AEAT ``RECIBIDAS_GASTOS`` ledger.

Fiscal interpretation belongs to ``aeat_received_expense_ledger_service``.
This module only filters its prepared rows by their frozen receipt date and
writes them to the official 42-column worksheet layout.
"""

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

from api.aeat_received_expense_ledger_service import (
    AEAT_RECEIVED_EXPENSE_LEDGER_COLUMN_KEYS,
    prepare_aeat_received_expense_ledger_rows,
)


AEAT_RECEIVED_EXPENSE_LEDGER_SHEET_NAME = "RECIBIDAS_GASTOS"
DATE_FORMAT = "dd/mm/yyyy"
MONEY_FORMAT = "#,##0.00"
PERCENT_FORMAT = "0.00"
PERIOD_QUARTERS = {"1T": 1, "2T": 2, "3T": 3, "4T": 4}

AEAT_RECEIVED_EXPENSE_HEADER_ROW_1 = [
    "Autoliquidación", "", "Actividad", "", "", "Tipo de factura",
    "Concepto de gasto", "Gasto deducible", "Fecha expedición",
    "Fecha operación", "Identificación de la factura", "", "Fecha recepción",
    "Número recepción", "", "Identificación del expedidor", "", "",
    "Nombre expedidor", "Clave operación", "Bien inversión",
    "Inversión sujeto pasivo", "Deducible periodo posterior", "",
    "", "Total factura", "Base imponible", "Tipo IVA", "Cuota IVA soportado",
    "Cuota deducible", "Tipo recargo equivalencia", "Cuota recargo equivalencia",
    "Pago", "", "", "", "Tipo retención IRPF", "Importe retenido IRPF",
    "Registro acuerdo facturación", "Inmueble", "", "Referencia externa",
]

AEAT_RECEIVED_EXPENSE_HEADER_ROW_2 = [
    "Ejercicio", "Periodo", "Código", "Tipo", "Grupo o Epígrafe del IAE", "",
    "", "", "", "", "Serie-Número", "Número-Final", "", "",
    "Número recepción final", "Tipo", "Código país", "Identificación", "", "",
    "", "", "", "Ejercicio deducción", "Periodo deducción", "", "", "",
    "", "", "", "", "Fecha", "Importe", "Medio utilizado",
    "Identificación medio utilizado", "", "", "", "Situación",
    "Referencia catastral", "",
]


class AeatReceivedExpenseLedgerWriteError(Exception):
    """Raised when the received-expense workbook cannot be written safely."""


@dataclass(frozen=True)
class AeatReceivedExpenseLedgerExportResult:
    output_path: str
    filename: str
    invoice_count: int
    row_count: int
    generated_at: datetime
    file_size: int


def generate_aeat_received_expense_ledger_workbook(supplier_invoices, *, year, period):
    """Build a standalone RECIBIDAS_GASTOS workbook without writing it."""
    prepared_rows = prepare_aeat_received_expense_ledger_rows(supplier_invoices)
    selected_rows = select_aeat_received_expense_ledger_rows(
        prepared_rows,
        year=year,
        period=period,
    )
    return _build_workbook(selected_rows), selected_rows


def export_aeat_received_expense_ledger(
    supplier_invoices,
    *,
    year,
    period,
    output_path,
    overwrite=False,
):
    """Write a standalone national received-expense workbook to ``output_path``."""
    path = _validated_output_path(output_path)
    if path.exists() and not overwrite:
        raise AeatReceivedExpenseLedgerWriteError(
            "El archivo AEAT de gastos ya existe."
        )

    workbook, selected_rows = generate_aeat_received_expense_ledger_workbook(
        supplier_invoices,
        year=year,
        period=period,
    )

    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
    except OSError as exc:
        raise AeatReceivedExpenseLedgerWriteError(
            "No se pudo escribir el libro AEAT de gastos."
        ) from exc

    return AeatReceivedExpenseLedgerExportResult(
        output_path=str(path),
        filename=path.name,
        invoice_count=len({row["reception_number"] for row in selected_rows}),
        row_count=len(selected_rows),
        generated_at=datetime.now(timezone.utc),
        file_size=path.stat().st_size,
    )


def select_aeat_received_expense_ledger_rows(rows, *, year, period):
    """Select cumulative quarters using only the frozen ``received_at`` row date."""
    normalized_year = _validated_year(year)
    target_quarter = _validated_period(period)
    return [
        row
        for row in rows
        if row["exercise"] == normalized_year
        and _quarter_number(row["period"]) <= target_quarter
    ]


def _build_workbook(rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = AEAT_RECEIVED_EXPENSE_LEDGER_SHEET_NAME
    worksheet.append(AEAT_RECEIVED_EXPENSE_HEADER_ROW_1)
    worksheet.append(AEAT_RECEIVED_EXPENSE_HEADER_ROW_2)

    for header_row in worksheet.iter_rows(min_row=1, max_row=2):
        for cell in header_row:
            cell.font = Font(bold=True)

    for row in rows:
        worksheet.append([row[key] for key in AEAT_RECEIVED_EXPENSE_LEDGER_COLUMN_KEYS])

    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A2:AP{worksheet.max_row}"
    _apply_formats(worksheet)
    _apply_column_widths(worksheet)
    return workbook


def _validated_output_path(output_path):
    if not output_path:
        raise AeatReceivedExpenseLedgerWriteError("La ruta de salida es obligatoria.")
    path = Path(output_path)
    if any(part == ".." for part in path.parts):
        raise AeatReceivedExpenseLedgerWriteError("La ruta de salida no es válida.")
    if path.suffix.lower() != ".xlsx":
        raise AeatReceivedExpenseLedgerWriteError(
            "La exportación AEAT debe tener extensión .xlsx."
        )
    if not path.name or path.name == ".xlsx":
        raise AeatReceivedExpenseLedgerWriteError(
            "El nombre del archivo de salida no es válido."
        )
    return path


def _validated_year(year):
    try:
        normalized_year = int(year)
    except (TypeError, ValueError) as exc:
        raise AeatReceivedExpenseLedgerWriteError("El ejercicio AEAT no es válido.") from exc
    if normalized_year < 1000 or normalized_year > 9999:
        raise AeatReceivedExpenseLedgerWriteError("El ejercicio AEAT no es válido.")
    return normalized_year


def _validated_period(period):
    if period not in PERIOD_QUARTERS:
        raise AeatReceivedExpenseLedgerWriteError("El período AEAT debe ser 1T, 2T, 3T o 4T.")
    return PERIOD_QUARTERS[period]


def _quarter_number(period):
    try:
        return PERIOD_QUARTERS[period]
    except KeyError as exc:
        raise AeatReceivedExpenseLedgerWriteError(
            "La fila preparada no tiene un período AEAT válido."
        ) from exc


def _apply_formats(worksheet):
    for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
        for index in (8, 9, 12):
            row[index].number_format = DATE_FORMAT
        for index in (7, 25, 26, 28, 29, 31, 33, 37):
            row[index].number_format = MONEY_FORMAT
        for index in (27, 30):
            row[index].number_format = PERCENT_FORMAT


def _apply_column_widths(worksheet):
    widths = {
        "A": 12, "B": 10, "C": 10, "D": 10, "E": 20, "F": 14,
        "G": 16, "H": 18, "I": 16, "J": 16, "K": 24, "L": 16,
        "M": 16, "N": 16, "O": 20, "P": 12, "Q": 14, "R": 18,
        "S": 34, "T": 16, "U": 16, "V": 20, "W": 22, "X": 18,
        "Y": 18, "Z": 16, "AA": 16, "AB": 12, "AC": 20, "AD": 18,
        "AE": 22, "AF": 22, "AG": 16, "AH": 16, "AI": 20, "AJ": 28,
        "AK": 18, "AL": 20, "AM": 24, "AN": 16, "AO": 24, "AP": 22,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
