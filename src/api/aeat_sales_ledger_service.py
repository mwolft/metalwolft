from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from api.invoice_snapshot_integrity import calculate_invoice_snapshot_hash


AEAT_SALES_LEDGER_SHEET_NAME = "EXPEDIDAS_INGRESOS"
SALE_ENTRY_TYPE = "sale"
ORDINARY_INVOICE_TYPE = "ordinary"
SUPPORTED_SCHEMA_VERSION = 1
SUPPORTED_CURRENCY = "EUR"
SUPPORTED_COUNTRY_CODE = "ES"
SUPPORTED_TAX_ID_TYPE = "4"
BUSINESS_ACTIVITY_CODE = "A"
BUSINESS_ACTIVITY_TYPE = "3"
IAE_CODE = "3141"
AEAT_INVOICE_TYPE = "F1"
AEAT_INCOME_CONCEPT = "I01"
AEAT_OPERATION_KEY = "1"
AEAT_OPERATION_QUALIFICATION = "S1"
DATE_FORMAT = "DD/MM/YYYY"
MONEY_FORMAT = '#,##0.00'
PERCENT_FORMAT = '0.00'

AEAT_HEADER_ROW_1 = [
    "Autoliquidación(11)",
    "",
    "Actividad(16)",
    "",
    "",
    "Tipo de Factura(9)",
    "Concepto de Ingreso(10) (17)",
    "Ingreso Computable(13) (17)",
    "Fecha Expedición(25)",
    "Fecha Operación(1)",
    "Identificación de la Factura",
    "",
    "",
    "NIF Destinatario(2)",
    "",
    "",
    "Nombre Destinatario",
    "Clave de Operación(6)(23)",
    "Calificación de la Operación(19) (21) (22) (23) (24) (30)",
    "Operación Exenta(20)",
    "Total Factura(37)",
    "Base Imponible",
    "Tipo de IVA(39)",
    "Cuota IVA Repercutida",
    "Tipo de Recargo Eq.",
    "Cuota Recargo Eq.(35)",
    "Cobro (Operación Criterio de Caja de IVA y/o artículo 7.2.1º de Reglamento del IRPF)",
    "",
    "",
    "",
    "Tipo Retención del IRPF(15) (17)",
    "Importe Retenido del IRPF(15) (17)",
    "Registro Acuerdo Facturación(18)",
    "Inmueble(40)",
    "",
    "Referencia Externa",
]

AEAT_HEADER_ROW_2 = [
    "Ejercicio",
    "Periodo",
    "Código",
    "Tipo",
    "Grupo o Epígrafe del IAE",
    "",
    "",
    "",
    "",
    "",
    "Serie",
    "Número",
    "Número-Final",
    "Tipo",
    "Código País",
    "Identificación",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "Fecha",
    "Importe",
    "Medio Utilizado",
    "Identificación Medio Utilizado",
    "",
    "",
    "",
    "Situación",
    "Referencia Catastral",
    "",
]


class AeatSalesLedgerError(Exception):
    """Base error for AEAT sales ledger exports."""


class AeatSalesLedgerValidationError(AeatSalesLedgerError):
    """Raised when entries cannot be represented in the AEAT ledger."""


class AeatSalesLedgerWriteError(AeatSalesLedgerError):
    """Raised when the AEAT ledger workbook cannot be written safely."""


@dataclass(frozen=True)
class AeatSalesLedgerExportResult:
    output_path: str
    filename: str
    row_count: int
    generated_at: datetime
    file_size: int


def export_aeat_sales_ledger(entries, *, output_path, overwrite=False):
    """Export sale AccountingEntry records to the AEAT EXPEDIDAS_INGRESOS layout."""
    generated_at = datetime.now(timezone.utc)
    path = _validated_output_path(output_path)
    prepared_entries = _prepared_entries(entries)

    if path.exists() and not overwrite:
        raise AeatSalesLedgerWriteError("El archivo AEAT de ingresos ya existe.")

    workbook = _build_workbook(prepared_entries)

    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
    except OSError as exc:
        raise AeatSalesLedgerWriteError("No se pudo escribir el libro AEAT de ingresos.") from exc

    return AeatSalesLedgerExportResult(
        output_path=str(path),
        filename=path.name,
        row_count=len(prepared_entries),
        generated_at=generated_at,
        file_size=path.stat().st_size,
    )


def generate_aeat_sales_ledger_workbook(entries):
    """Build the workbook without saving it, useful for tests and future callers."""
    return _build_workbook(_prepared_entries(entries))


def _build_workbook(prepared_entries):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = AEAT_SALES_LEDGER_SHEET_NAME
    worksheet.append(AEAT_HEADER_ROW_1)
    worksheet.append(AEAT_HEADER_ROW_2)

    for row in worksheet.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.font = Font(bold=True)

    for entry in prepared_entries:
        worksheet.append(_aeat_row(entry))

    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A2:AJ{worksheet.max_row}"
    _apply_formats(worksheet)
    _apply_column_widths(worksheet)
    return workbook


def _validated_output_path(output_path):
    if not output_path:
        raise AeatSalesLedgerValidationError("La ruta de salida es obligatoria.")

    path = Path(output_path)
    if any(part == ".." for part in path.parts):
        raise AeatSalesLedgerValidationError("La ruta de salida no es valida.")
    if path.suffix.lower() != ".xlsx":
        raise AeatSalesLedgerValidationError("La exportacion AEAT debe tener extension .xlsx.")
    if not path.name or path.name == ".xlsx":
        raise AeatSalesLedgerValidationError("El nombre del archivo de salida no es valido.")
    return path


def _prepared_entries(entries):
    if entries is None:
        raise AeatSalesLedgerValidationError("Debe indicarse al menos un registro contable.")

    prepared = [_prepared_entry(entry) for entry in list(entries)]
    if not prepared:
        raise AeatSalesLedgerValidationError("Debe indicarse al menos un registro contable.")

    return sorted(
        prepared,
        key=lambda entry: (
            entry["issue_date"],
            entry["invoice_number"],
            entry["id"],
        ),
    )


def _prepared_entry(entry):
    if _required_text(getattr(entry, "entry_type", None), "entry_type") != SALE_ENTRY_TYPE:
        raise AeatSalesLedgerValidationError("Solo se pueden exportar registros contables de venta.")

    entry_currency = _required_text(getattr(entry, "currency", None), "currency").upper()
    if entry_currency != SUPPORTED_CURRENCY:
        raise AeatSalesLedgerValidationError("Moneda no soportada para el libro AEAT.")

    invoice = getattr(entry, "invoice", None)
    if invoice is None:
        raise AeatSalesLedgerValidationError("La factura emitida es obligatoria para el libro AEAT.")

    snapshot = _validated_snapshot(invoice)
    _validate_snapshot_hash(invoice, snapshot)
    _validate_ordinary_invoice(invoice, snapshot)

    invoice_number = _required_text(getattr(entry, "invoice_number", None), "invoice_number")
    if invoice_number != _required_text(getattr(invoice, "invoice_number", None), "invoice.invoice_number"):
        raise AeatSalesLedgerValidationError("La proyeccion contable no coincide con la factura emitida.")

    snapshot_currency = _required_text(snapshot["operation"].get("currency"), "operation.currency").upper()
    if snapshot_currency != SUPPORTED_CURRENCY or snapshot_currency != entry_currency:
        raise AeatSalesLedgerValidationError("La moneda fiscal no coincide con la proyeccion contable.")

    issue_date = _snapshot_date(snapshot["operation"].get("issue_date"), "operation.issue_date")
    operation_date = _operation_date(snapshot, issue_date)
    customer = snapshot["customer"]
    totals = snapshot["totals"]

    tax_base = _money(totals.get("tax_base"), "totals.tax_base")
    tax_amount = _money(totals.get("tax_amount"), "totals.tax_amount")
    total_amount = _money(totals.get("total_amount"), "totals.total_amount")
    _validate_entry_totals(entry, tax_base, tax_amount, total_amount)

    return {
        "id": _entry_id(entry),
        "exercise": issue_date.year,
        "period": _quarter(issue_date),
        "invoice_number": invoice_number,
        "issue_date": issue_date,
        "operation_date": operation_date,
        "customer_name": _required_text(customer.get("legal_name"), "customer.legal_name"),
        "customer_tax_id": _required_text(customer.get("tax_id"), "customer.tax_id"),
        "country_code": _country_code(customer),
        "tax_id_type": SUPPORTED_TAX_ID_TYPE,
        "tax_base": tax_base,
        "tax_amount": tax_amount,
        "total_amount": total_amount,
        "tax_rate": _tax_rate(snapshot, tax_base, tax_amount),
        "reference": _external_reference(snapshot),
    }


def _validated_snapshot(invoice):
    if not getattr(invoice, "issued_at", None):
        raise AeatSalesLedgerValidationError("La factura debe estar emitida.")
    if not getattr(invoice, "invoice_number", None):
        raise AeatSalesLedgerValidationError("La factura emitida debe tener numero.")

    snapshot = getattr(invoice, "invoice_snapshot", None)
    if not isinstance(snapshot, dict):
        raise AeatSalesLedgerValidationError("La factura no tiene snapshot fiscal.")
    if snapshot.get("schema_version") != SUPPORTED_SCHEMA_VERSION:
        raise AeatSalesLedgerValidationError("Version de snapshot fiscal no soportada.")

    for block in ("customer", "operation", "totals"):
        if not isinstance(snapshot.get(block), dict):
            raise AeatSalesLedgerValidationError(f"Bloque de snapshot obligatorio ausente: {block}.")

    return snapshot


def _validate_snapshot_hash(invoice, snapshot):
    stored_hash = getattr(invoice, "invoice_snapshot_hash", None)
    if not stored_hash:
        raise AeatSalesLedgerValidationError("Hash de snapshot fiscal ausente.")
    if calculate_invoice_snapshot_hash(snapshot) != stored_hash:
        raise AeatSalesLedgerValidationError("La integridad del snapshot fiscal no coincide.")


def _validate_ordinary_invoice(invoice, snapshot):
    model_type = getattr(invoice, "invoice_type", None)
    snapshot_type = snapshot["operation"].get("invoice_type")
    invoice_type = snapshot_type or model_type
    if invoice_type != ORDINARY_INVOICE_TYPE:
        raise AeatSalesLedgerValidationError("Solo se soportan facturas ordinarias en el libro AEAT v1.")


def _validate_entry_totals(entry, tax_base, tax_amount, total_amount):
    expected = {
        "taxable_base": tax_base,
        "vat_amount": tax_amount,
        "total_amount": total_amount,
    }
    for field, amount in expected.items():
        if _money(getattr(entry, field, None), field) != amount:
            raise AeatSalesLedgerValidationError("La proyeccion contable no coincide con el snapshot fiscal.")


def _country_code(customer):
    country_code = customer.get("country_code") or SUPPORTED_COUNTRY_CODE
    country_code = str(country_code).strip().upper()
    if country_code != SUPPORTED_COUNTRY_CODE:
        raise AeatSalesLedgerValidationError("Solo se soportan destinatarios espanoles en el libro AEAT v1.")
    return country_code


def _tax_rate(snapshot, tax_base, tax_amount):
    line_rates = set()
    lines = snapshot.get("lines")
    if isinstance(lines, list):
        for index, line in enumerate(lines, start=1):
            if not isinstance(line, dict):
                raise AeatSalesLedgerValidationError(f"Linea fiscal invalida: {index}.")
            rate = line.get("tax_rate")
            if rate not in (None, ""):
                line_rates.add(_money(rate, f"lines.{index}.tax_rate"))

    if len(line_rates) > 1:
        raise AeatSalesLedgerValidationError("El libro AEAT v1 no soporta facturas con varios tipos de IVA.")
    if len(line_rates) == 1:
        return next(iter(line_rates))

    if tax_base <= Decimal("0.00"):
        raise AeatSalesLedgerValidationError("No se puede derivar el tipo de IVA con base imponible cero.")
    return ((tax_amount / tax_base) * Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _operation_date(snapshot, issue_date):
    operation = snapshot["operation"]
    raw_operation_date = operation.get("operation_date") or operation.get("service_date") or issue_date
    return _snapshot_date(raw_operation_date, "operation.operation_date")


def _snapshot_date(value, field):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value:
        try:
            return datetime.fromisoformat(value).date()
        except ValueError as exc:
            raise AeatSalesLedgerValidationError(f"Fecha invalida en {field}.") from exc
    raise AeatSalesLedgerValidationError(f"Fecha obligatoria ausente: {field}.")


def _money(value, field):
    try:
        amount = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise AeatSalesLedgerValidationError(f"Importe no valido en {field}.") from exc
    if not amount.is_finite():
        raise AeatSalesLedgerValidationError(f"Importe no valido en {field}.")
    return amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _required_text(value, field):
    text = _optional_text(value)
    if not text:
        raise AeatSalesLedgerValidationError(f"Campo obligatorio ausente: {field}.")
    return text


def _optional_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _entry_id(entry):
    entry_id = getattr(entry, "id", None)
    return int(entry_id or 0)


def _quarter(issue_date):
    return f"{((issue_date.month - 1) // 3) + 1}T"


def _external_reference(snapshot):
    references = snapshot.get("references")
    if not isinstance(references, dict):
        return None
    order_id = references.get("order_id")
    return f"order:{order_id}" if order_id not in (None, "") else None


def _aeat_row(entry):
    return [
        entry["exercise"],
        entry["period"],
        BUSINESS_ACTIVITY_CODE,
        BUSINESS_ACTIVITY_TYPE,
        IAE_CODE,
        AEAT_INVOICE_TYPE,
        AEAT_INCOME_CONCEPT,
        entry["tax_base"],
        entry["issue_date"],
        entry["operation_date"],
        None,
        entry["invoice_number"],
        None,
        entry["tax_id_type"],
        entry["country_code"],
        entry["customer_tax_id"],
        entry["customer_name"],
        AEAT_OPERATION_KEY,
        AEAT_OPERATION_QUALIFICATION,
        None,
        entry["total_amount"],
        entry["tax_base"],
        entry["tax_rate"],
        entry["tax_amount"],
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        entry["reference"],
    ]


def _apply_formats(worksheet):
    for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
        row[8].number_format = DATE_FORMAT
        row[9].number_format = DATE_FORMAT
        for index in (7, 20, 21, 23):
            row[index].number_format = MONEY_FORMAT
        row[22].number_format = PERCENT_FORMAT


def _apply_column_widths(worksheet):
    widths = {
        "A": 12,
        "B": 10,
        "C": 10,
        "D": 10,
        "E": 18,
        "F": 14,
        "G": 18,
        "H": 18,
        "I": 16,
        "J": 16,
        "K": 12,
        "L": 22,
        "M": 14,
        "N": 12,
        "O": 12,
        "P": 18,
        "Q": 30,
        "R": 16,
        "S": 18,
        "U": 16,
        "V": 16,
        "W": 12,
        "X": 18,
        "AJ": 18,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
